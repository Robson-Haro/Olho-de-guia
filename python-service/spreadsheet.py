"""Geração da planilha Excel do Eureka em memória."""

from __future__ import annotations

from datetime import datetime, timezone
from io import BytesIO
from typing import Any, Iterable

from openpyxl import Workbook
from openpyxl.styles import Alignment, Font, PatternFill
from openpyxl.utils import get_column_letter


HEADERS = [
    "Ranking",
    "Aderência",
    "Confiança das evidências",
    "Nome",
    "Cargo atual",
    "Empresa",
    "Cidade",
    "Estado",
    "Competências identificadas",
    "Motivo da aderência",
    "Resumo público",
    "LinkedIn",
    "Fonte",
]


def safe_cell_value(value: Any) -> Any:
    """Neutraliza fórmulas vindas de textos públicos antes de gravar no Excel."""
    if not isinstance(value, str):
        return value
    stripped = value.lstrip()
    return f"'{value}" if stripped.startswith(("=", "+", "-", "@")) else value


def create_candidate_workbook(job: dict[str, Any], candidates: Iterable[dict[str, Any]]) -> bytes:
    workbook = Workbook()
    sheet = workbook.active
    sheet.title = "Candidatos"
    sheet.freeze_panes = "A2"
    sheet.auto_filter.ref = f"A1:M1"

    header_fill = PatternFill("solid", fgColor="07152B")
    accent_fill = PatternFill("solid", fgColor="006EB8")
    header_font = Font(color="FFFFFF", bold=True)
    link_font = Font(color="0563C1", underline="single")

    for column, header in enumerate(HEADERS, start=1):
        cell = sheet.cell(row=1, column=column, value=header)
        cell.fill = header_fill if column not in (1, 2) else accent_fill
        cell.font = header_font
        cell.alignment = Alignment(horizontal="center", vertical="center", wrap_text=True)
    sheet.row_dimensions[1].height = 34

    for row_index, candidate in enumerate(candidates, start=2):
        matched_skills = candidate.get("matchedSkills")
        if not isinstance(matched_skills, list):
            matched_skills = []
        values = [
            candidate.get("rank") or row_index - 1,
            (float(candidate.get("compatibility") or 0) / 100),
            candidate.get("evidenceLabel") or "não calculada",
            candidate.get("name") or "",
            candidate.get("title") or "",
            candidate.get("company") or "",
            candidate.get("city") or "",
            candidate.get("state") or "",
            ", ".join(str(skill) for skill in matched_skills),
            candidate.get("matchReason") or "",
            candidate.get("summary") or "",
            candidate.get("profileUrl") or "",
            candidate.get("source") or "Google via Serper",
        ]
        for column, value in enumerate(values, start=1):
            cell = sheet.cell(row=row_index, column=column, value=safe_cell_value(value))
            cell.alignment = Alignment(vertical="top", wrap_text=True)
        sheet.cell(row=row_index, column=2).number_format = "0%"
        link = str(candidate.get("profileUrl") or "")
        if link:
            link_cell = sheet.cell(row=row_index, column=12)
            link_cell.hyperlink = link
            link_cell.font = link_font

    widths = [10, 12, 22, 28, 32, 25, 20, 10, 34, 54, 58, 45, 20]
    for index, width in enumerate(widths, start=1):
        sheet.column_dimensions[get_column_letter(index)].width = width

    details = workbook.create_sheet("Detalhes da busca")
    details.append(["Campo", "Informação"])
    details.append(["Vaga", safe_cell_value(job.get("title") or "")])
    details.append(["Cidade principal", safe_cell_value(job.get("city") or "Brasil inteiro")])
    details.append(["Cidade adicional", safe_cell_value(job.get("additionalCity") or "")])
    details.append(["Busca nacional", "Sim" if job.get("nationwide") is True else "Não"])
    details.append(["Descrição da vaga", safe_cell_value(job.get("description") or "")])
    details.append(["Gerado em UTC", datetime.now(timezone.utc).strftime("%d/%m/%Y %H:%M")])
    details.append(["Critério", "Ranking profissional explicável; decisão final deve ser humana."])
    for cell in details[1]:
        cell.fill = header_fill
        cell.font = header_font
    details.column_dimensions["A"].width = 24
    details.column_dimensions["B"].width = 110
    for row in details.iter_rows():
        for cell in row:
            cell.alignment = Alignment(vertical="top", wrap_text=True)

    buffer = BytesIO()
    workbook.save(buffer)
    return buffer.getvalue()
