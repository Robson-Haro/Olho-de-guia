from io import BytesIO
import unittest

from openpyxl import load_workbook

from engine import analyze_job, rank_candidates
from spreadsheet import create_candidate_workbook


class TalentEngineTests(unittest.TestCase):
    def setUp(self):
        self.job = {
            "title": "Analista de Administração de Pessoal",
            "city": "São Paulo",
            "additionalCity": "",
            "description": "Responsável por folha de pagamento, admissão, rescisão, Excel e legislação trabalhista.",
            "keywords": ["folha de pagamento", "Excel"],
            "nationwide": True,
        }

    def test_expands_titles_in_three_languages(self):
        intelligence = analyze_job(self.job)
        normalized = " | ".join(intelligence.equivalent_titles).lower()
        self.assertIn("payroll analyst", normalized)
        self.assertIn("analista de nómina", normalized)
        self.assertIn("analista de departamento pessoal", normalized)

    def test_multilingual_equivalent_outranks_unrelated_profile(self):
        candidates = [
            {
                "id": "1", "name": "Ana", "title": "Payroll Analyst", "summary": "Payroll, Excel and labor law",
                "city": "São Paulo", "state": "SP", "profileUrl": "https://www.linkedin.com/in/ana",
            },
            {
                "id": "2", "name": "Bruno", "title": "Marketing Analyst", "summary": "Digital campaigns and branding",
                "city": "São Paulo", "state": "SP", "profileUrl": "https://www.linkedin.com/in/bruno",
            },
        ]
        _, ranked = rank_candidates(self.job, candidates)
        # Perfis de outra família profissional não são exibidos para completar
        # artificialmente a quantidade solicitada.
        self.assertEqual("Ana", ranked[0]["name"])
        self.assertEqual(["Ana"], [candidate["name"] for candidate in ranked])
        self.assertEqual("Python 3 · motor multilíngue", ranked[0]["rankingEngine"])

    def test_executive_hr_search_rejects_unrelated_and_junior_profiles(self):
        job = {
            **self.job,
            "title": "Gerente Executivo de Recursos Humanos",
            "description": (
                "Liderança integral de Recursos Humanos no Brasil, HR Business Partner estratégico, "
                "operação industrial, relações sindicais, engajamento, turnover, P&L de RH e "
                "gestão da força de trabalho."
            ),
            "keywords": [],
        }
        candidates = [
            {"name": "Diretora RH", "title": "HR Director", "company": "Indústria Alfa",
             "summary": "Strategic HRBP, plant HR, labor relations, engagement and workforce planning"},
            {"name": "Cleiton Ito", "title": "Analista de Recursos Humanos", "company": "Unilever",
             "summary": "Analista de Recursos Humanos em São Paulo"},
            {"name": "Christian", "title": "Senior Account Manager", "summary": "Business development"},
            {"name": "Diogo", "title": "Coordenador de Suporte", "summary": "KPIs, Power BI e Excel"},
            {"name": "Andreia", "title": "Trader", "summary": "Global commercial strategies"},
            {"name": "Hellen", "title": "Fraud Analyst", "summary": "Fast-paced environments"},
            {"name": "Claudia", "title": "Executive Assistant and Office Manager", "summary": "Office management"},
        ]
        intelligence, ranked = rank_candidates(job, candidates)
        self.assertEqual("human_resources", intelligence.family)
        self.assertEqual(["Diretora RH"], [candidate["name"] for candidate in ranked])
        self.assertIn("HR Business Partner estratégico", intelligence.skills)
        self.assertIn("Relações trabalhistas e sindicais", intelligence.skills)
        self.assertIn("RH em operação industrial", intelligence.skills)

    def test_process_standardization_expands_titles_in_three_languages(self):
        job = {
            **self.job,
            "title": "Gerente de Padronização de Processos",
            "description": "Gestão, governança e padronização de processos industriais.",
            "keywords": ["Couro", "Curtume"],
        }
        intelligence = analyze_job(job)
        normalized = " | ".join(intelligence.equivalent_titles).lower()
        self.assertEqual("process_excellence", intelligence.family)
        self.assertIn("process standardization manager", normalized)
        self.assertIn("gerente de estandarización de procesos", normalized)
        self.assertEqual(["Couro / Leather", "Curtume / Tannery"], [
            concept.label for concept in intelligence.required_keywords
        ])

    def test_required_keywords_classify_profiles_without_public_evidence(self):
        job = {
            **self.job,
            "title": "Gerente de Padronização de Processos",
            "description": "Padronização, melhoria contínua e processos industriais.",
            "keywords": ["Couro", "Curtume"],
        }
        candidates = [
            {
                "id": "1", "name": "Ana", "title": "Process Standardization Manager",
                "summary": "Process governance and operational excellence in the leather and tannery industry",
                "city": "Franca", "state": "SP", "profileUrl": "https://www.linkedin.com/in/ana",
            },
            {
                "id": "2", "name": "Bruno", "title": "Gerente de Processos",
                "summary": "Padronização de processos no setor bancário",
                "city": "São Paulo", "state": "SP", "profileUrl": "https://www.linkedin.com/in/bruno",
            },
            {
                "id": "3", "name": "Carlos", "title": "Gerente de Procesos",
                "summary": "Excelencia operacional en curtiembre y curtido de cuero",
                "city": "Buenos Aires", "state": "", "profileUrl": "https://www.linkedin.com/in/carlos",
            },
            {
                "id": "4", "name": "Diana", "title": "Continuous Improvement Manager",
                "summary": "Lean transformation and process governance for tannery operations",
                "city": "Franca", "state": "SP", "profileUrl": "https://www.linkedin.com/in/diana",
            },
            {
                "id": "5", "name": "Eduardo", "title": "Gerente Comercial",
                "summary": "Vendas de artigos de couro no varejo",
                "city": "São Paulo", "state": "SP", "profileUrl": "https://www.linkedin.com/in/eduardo",
            },
        ]
        _, ranked = rank_candidates(job, candidates)
        # Quem evidencia todos os critérios obrigatórios ocupa as primeiras
        # posições; perfis de outra família profissional não são exibidos.
        top = [candidate["name"] for candidate in ranked[:3]]
        self.assertEqual(["Ana", "Carlos", "Diana"], sorted(top))
        self.assertTrue(all(candidate["tier"] == "A" for candidate in ranked[:3]))
        self.assertTrue(all(not candidate["missingRequiredKeywords"] for candidate in ranked[:3]))
        self.assertTrue(all(len(candidate["matchedRequiredKeywords"]) == 2 for candidate in ranked[:3]))
        self.assertNotIn("Eduardo", [candidate["name"] for candidate in ranked])

    def test_segment_company_evidence_improves_ranking(self):
        job = {
            **self.job,
            "title": "Analista Financeiro",
            "description": "Planejamento financeiro, budget, forecast e análise de resultados.",
            "marketSegment": "beef_processing",
            "mappedCompanies": ["Minerva Foods", "JBS", "Marfrig"],
        }
        candidates = [
            {
                "id": "1", "name": "Ana", "title": "Analista Financeiro",
                "company": "Minerva Foods", "summary": "Budget, forecast e Excel",
                "city": "São Paulo", "state": "SP", "profileUrl": "https://www.linkedin.com/in/ana",
            },
            {
                "id": "2", "name": "Bruno", "title": "Analista Financeiro",
                "company": "Empresa não mapeada", "summary": "Budget, forecast e Excel",
                "city": "São Paulo", "state": "SP", "profileUrl": "https://www.linkedin.com/in/bruno",
            },
        ]
        _, ranked = rank_candidates(job, candidates)
        self.assertEqual("Ana", ranked[0]["name"])
        self.assertEqual(10, ranked[0]["scoreBreakdown"]["segmento"])
        self.assertEqual(2, ranked[1]["scoreBreakdown"]["segmento"])
        self.assertIn("empresa do segmento", ranked[0]["matchReason"])

    def test_manager_outranks_analyst_in_the_same_professional_family(self):
        job = {
            **self.job,
            "title": "Gerente de Padronização de Processos",
            "description": "Governança e padronização de processos industriais em curtume.",
            "keywords": ["Couro e Curtume"],
        }
        candidates = [
            {
                "id": "1", "name": "Gerente", "title": "Process Standardization Manager",
                "summary": "Process governance for leather and tannery operations",
                "city": "São Paulo", "state": "SP", "country": "Brasil",
                "profileUrl": "https://www.linkedin.com/in/gerente",
            },
            {
                "id": "2", "name": "Analista", "title": "Analista de Padronização de Processos",
                "summary": "Governança de processos de couro e curtume",
                "city": "São Paulo", "state": "SP", "country": "Brasil",
                "profileUrl": "https://www.linkedin.com/in/analista",
            },
        ]
        _, ranked = rank_candidates(job, candidates)
        self.assertEqual("Gerente", ranked[0]["name"])
        self.assertEqual(["Gerente"], [candidate["name"] for candidate in ranked])
        self.assertIn("senioridade compatível", ranked[0]["matchReason"])

    def test_skill_explanation_never_reports_more_matches_than_the_denominator(self):
        job = {
            **self.job,
            "title": "Gerente de Padronização de Processos",
            "description": "Gestão e padronização de processos industriais em couro e curtume.",
            "keywords": ["Couro e Curtume"],
        }
        _, ranked = rank_candidates(job, [{
            "id": "1", "name": "Aderente", "title": "Gerente de Processos Industriais",
            "company": "Curtume Alfa", "summary": "Padronização de processos de couro em curtume",
            "city": "São Paulo", "state": "SP", "country": "Brasil",
            "profileUrl": "https://www.linkedin.com/in/aderente",
        }])
        match = ranked[0]
        ratio = next(part for part in match["matchReason"].split(" · ") if "competência(s)" in part)
        numerator, denominator = (int(value) for value in ratio.split()[0].split("/"))
        self.assertLessEqual(numerator, denominator)
        self.assertEqual("A", match["tier"])
        self.assertIn("completa", match["tierLabel"])
        self.assertIn("evidencia", match["scoreBreakdown"])

    def test_generic_terms_and_explicit_keyword_aliases_do_not_duplicate_skills(self):
        job = {
            **self.job,
            "title": "Gerente de Padronização de Processos",
            "description": "Gestão de equipes, processos, resultados e padronização de processos em couro e curtume.",
            "keywords": ["Couro e Curtume"],
        }
        intelligence = analyze_job(job)
        normalized = [skill.lower() for skill in intelligence.skills]
        self.assertNotIn("gestao", normalized)
        self.assertNotIn("processos", normalized)
        self.assertNotIn("couro", normalized)
        self.assertNotIn("curtume", normalized)
        self.assertEqual(1, normalized.count("couro / leather"))
        self.assertEqual(1, normalized.count("curtume / tannery"))

    def test_composite_keyword_field_becomes_two_required_concepts(self):
        job = {**self.job, "keywords": ["Couro e Curtume"]}
        intelligence = analyze_job(job)
        self.assertEqual(["Couro / Leather", "Curtume / Tannery"], [
            concept.label for concept in intelligence.required_keywords
        ])

    def test_sensitive_traits_never_become_ranking_skills(self):
        job = {**self.job, "keywords": ["Excel", "idade", "PcD", "gênero"]}
        intelligence = analyze_job(job)
        normalized_skills = " ".join(intelligence.skills).lower()
        self.assertIn("excel", normalized_skills)
        self.assertNotIn("idade", normalized_skills)
        self.assertNotIn("pcd", normalized_skills)
        self.assertNotIn("gênero", normalized_skills)

    def test_multicity_geography_prioritizes_selected_city_and_country(self):
        job = {
            **self.job,
            "countryCode": "MX",
            "country": "México",
            "subdivision": "Nuevo León",
            "cities": ["Monterrey", "Guadalupe"],
            "countrywide": False,
            "nationwide": False,
        }
        candidates = [
            {
                "id": "1", "name": "María", "title": "Payroll Analyst",
                "summary": "Payroll, Excel and labor law in Monterrey, Nuevo León, México",
                "city": "Monterrey", "state": "Nuevo León", "country": "México",
                "profileUrl": "https://www.linkedin.com/in/maria",
            },
            {
                "id": "2", "name": "Laura", "title": "Payroll Analyst",
                "summary": "Payroll, Excel and labor law in Bogotá, Colombia",
                "city": "Bogotá", "state": "Bogotá D.C.", "country": "Colombia",
                "profileUrl": "https://www.linkedin.com/in/laura",
            },
        ]
        _, ranked = rank_candidates(job, candidates)
        self.assertEqual("María", ranked[0]["name"])
        self.assertIn("cidade selecionada confirmada", ranked[0]["matchReason"])

    def test_countrywide_search_requires_public_country_evidence_for_full_location_score(self):
        job = {
            **self.job,
            "countryCode": "MX",
            "country": "México",
            "subdivision": "",
            "cities": [],
            "countrywide": True,
            "nationwide": True,
        }
        candidates = [
            {
                "id": "1", "name": "México", "title": "Payroll Analyst",
                "summary": "Payroll and Excel in México", "country": "México",
                "profileUrl": "https://www.linkedin.com/in/mexico",
            },
            {
                "id": "2", "name": "Outro país", "title": "Payroll Analyst",
                "summary": "Payroll and Excel", "country": "",
                "profileUrl": "https://www.linkedin.com/in/outro",
            },
        ]
        _, ranked = rank_candidates(job, candidates)
        self.assertEqual("México", ranked[0]["name"])
        self.assertGreater(ranked[0]["compatibility"], ranked[1]["compatibility"])

    def test_export_creates_real_xlsx_with_hyperlink(self):
        _, ranked = rank_candidates(self.job, [{
            "id": "1", "name": "Ana", "title": "Payroll Analyst", "summary": "Payroll and Excel",
            "city": "São Paulo", "state": "SP", "profileUrl": "https://www.linkedin.com/in/ana",
        }])
        binary = create_candidate_workbook(self.job, ranked)
        workbook = load_workbook(BytesIO(binary))
        sheet = workbook["Candidatos"]
        self.assertEqual("Nome", sheet["D1"].value)
        self.assertEqual("Ana", sheet["D2"].value)
        self.assertEqual("https://www.linkedin.com/in/ana", sheet["L2"].hyperlink.target)

    def test_export_neutralizes_excel_formulas(self):
        _, ranked = rank_candidates(self.job, [{
            "id": "1", "name": "=HYPERLINK(\"bad\")", "title": "Payroll Analyst", "summary": "Payroll and Excel",
            "city": "São Paulo", "state": "SP", "profileUrl": "https://www.linkedin.com/in/safe",
        }])
        binary = create_candidate_workbook(self.job, ranked)
        workbook = load_workbook(BytesIO(binary), data_only=False)
        self.assertTrue(str(workbook["Candidatos"]["D2"].value).startswith("'="))


if __name__ == "__main__":
    unittest.main()
